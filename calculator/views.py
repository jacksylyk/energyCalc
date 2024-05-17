import json
from datetime import timezone

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.db.models import Q
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.urls.base import reverse
from django.views.generic import ListView, CreateView, DetailView, TemplateView
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.writer.excel import save_virtual_workbook

from calculator.models import Client, Contact, Invoice
import openpyxl
from openpyxl.utils import get_column_letter


# Create your views here.
@login_required
def index(request):
    return render(request, 'calculator/home.html')


def create_contacts_from_json(client_instance, json_data):
    contacts_array = json.loads(json_data)
    for contact_data in contacts_array:
        contact = Contact.objects.create(client=client_instance, contact=contact_data['phone'])


def format_date(date):
    if date:
        return date.strftime('%d.%m.%Y')
    else:
        return None


class ClientCreateView(LoginRequiredMixin, CreateView):
    model = Client
    fields = '__all__'
    template_name = "calculator/create_client.html"
    success_url = reverse_lazy('clients')

    def form_valid(self, form):
        # Save client and contacts
        form.instance.user = self.request.user
        return self.save_with_contacts(form)

    def save_with_contacts(self, form):
        contacts_array = self.request.POST.get('contacts_array', None)
        if contacts_array:
            client_instance = form.save(commit=False)
            client_instance.save()
            create_contacts_from_json(client_instance, contacts_array)
        else:
            form.save()
        return super().form_valid(form)


class ClientListView(LoginRequiredMixin, ListView):
    model = Client
    context_object_name = 'clients'

    def get_queryset(self):
        queryset = Client.objects.all()
        search_query = self.request.GET.get('q')
        if search_query:
            queryset = queryset.filter(
                Q(bin_number__icontains=search_query) | Q(contract_number__icontains=search_query))
        return queryset

    template_name = 'calculator/clients.html'


class ClientDetailView(DetailView):
    model = Client
    template_name = 'calculator/client_detail.html'
    context_object_name = 'client'

    def get_queryset(self):
        return super().get_queryset().prefetch_related('invoices')


def export_checked_clients(request):
    if request.method == 'POST':
        selected_client_ids = request.POST.getlist('client_id')
        selected_clients = Client.objects.filter(pk__in=selected_client_ids)

        print(selected_clients)
        return json.dumps(selected_clients)


class CheckedClientsExportView(View):
    def post(self, request):
        selected_client_ids = request.POST.getlist('client_id')
        clients = Client.objects.filter(pk__in=selected_client_ids)
        # Create a new workbook
        wb = Workbook()
        ws = wb.active

        # Define the headers
        headers = [
            'Наименование клиента', 'Местоположение ПКУ', 'Информация по ПКУ', 'Дата акта', 'Напряжение (U)',
            'Начало показании', 'Конец показании', 'Расход', "Создан"
        ]

        # Write the headers to the worksheet
        ws.append(headers)

        # Make the header cells bold and apply border
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True)
                cell.border = Border(bottom=Side(style='thin'))
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Write client data to the worksheet
        for client in clients:
            invoices = Invoice.objects.filter(client=client)
            for invoice in invoices:
                ws.append([
                    client.name,
                    client.location,
                    client.information,
                    format_date(client.contract_date),
                    invoice.voltage,
                    invoice.start,
                    invoice.end,
                    invoice.consumption,
                    format_date(invoice.created_at)
                ])

        # Apply border to all cells in the table
        for row in ws.iter_rows(min_row=1, max_row=len(clients) + 2, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'),
                                     right=Side(style='thin'))
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        # Adjust column widths for better readability
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 1) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        # Save the workbook to a virtual file
        virtual_workbook = save_virtual_workbook(wb)

        # Return the Excel file as a response
        response = HttpResponse(
            virtual_workbook,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=clients_invoices.xlsx'
        return response


class ClientExportView(View):
    def get(self, request, client_id):
        # Get the client by ID
        try:
            client = Client.objects.get(id=client_id)
        except Client.DoesNotExist:
            return HttpResponse("Client not found", status=404)

        invoices = Invoice.objects.filter(client=client)

        excel_data = self.export_invoices_to_excel(invoices)

        response = HttpResponse(
            excel_data,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={client.name}_invoices.xlsx'
        return response

    def export_invoices_to_excel(self, invoices):
        # Create a new workbook
        wb = Workbook()
        ws = wb.active

        # Define the headers
        headers = [
            'Наименование клиента', 'Местоположение ПКУ', 'Информация по ПКУ', 'Дата акта', 'Напряжение (U)',
            'Начало показании', 'Конец показании', 'Расход', "Создан"
        ]

        # Write the headers to the worksheet
        ws.append(headers)

        # Make the header cells bold and apply border
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True)
                cell.border = Border(bottom=Side(style='thin'))
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Write the invoice data to the worksheet
        for invoice in invoices:
            ws.append([
                invoice.client.name,
                invoice.client.location,
                invoice.client.information,
                format_date(invoice.client.contract_date),
                invoice.voltage,
                invoice.start,
                invoice.end,
                invoice.consumption,
                format_date(invoice.created_at)
            ])

        # Apply border to all cells in the table
        for row in ws.iter_rows(min_row=1, max_row=len(invoices) + 1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'),
                                     right=Side(style='thin'))
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Calculate the total for the "Расход" column
        total_row = ['Total']
        for _ in range(len(headers) - 3):  # Skip the first empty cell
            total_row.append('')

        # Calculate the total for the "Расход" column
        total_consumption = sum(invoice.consumption for invoice in invoices)
        total_row.append(total_consumption)
        ws.append(total_row)

        # Apply border to the total row
        for row in ws.iter_rows(min_row=len(invoices) + 2, max_row=len(invoices) + 2, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'),
                                     right=Side(style='thin'))

        # Adjust column widths for better readability
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        # Save the workbook to a virtual file
        virtual_workbook = save_virtual_workbook(wb)
        return virtual_workbook


class InvoiceCreateView(LoginRequiredMixin, CreateView):
    model = Invoice
    fields = '__all__'
    template_name = 'calculator/create_invoice.html'

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        if self.object:  # Check if self.object exists
            pk = self.object.pk
            return reverse_lazy('invoice_detail', kwargs={'pk': pk})
        else:
            # Provide a fallback URL if self.object is not defined
            return reverse_lazy('index')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        client_id = self.request.GET.get('client_id')
        if client_id:
            kwargs['initial'] = {'client': client_id}
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['clients'] = Client.objects.all()
        return context


class InvoiceDetailView(LoginRequiredMixin, DetailView):
    model = Invoice
    template_name = 'calculator/invoice_detail.html'
    context_object_name = 'invoice'

    def get_object(self, queryset=None):
        pk = self.kwargs.get('pk')
        return get_object_or_404(Invoice, pk=pk)


